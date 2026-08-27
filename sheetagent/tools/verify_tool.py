"""Tool: independently verify that both imports actually landed.

The agent is not allowed to declare success on the strength of earlier tool
output alone - this re-reads the workbook from disk and the live Google Sheet
and compares both against the CSV.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from ..registry import ToolContext, tool
from .excel_tool import read_csv

log = logging.getLogger("sheetagent.tools.verify")


def _read_workbook(path: Path) -> tuple[list[str], int]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        header, body = read_csv(path)
        return header, len(body)
    if suffix == ".ods":
        from odf.opendocument import load
        from odf.table import Table, TableRow
        doc = load(str(path))
        table = doc.spreadsheet.getElementsByType(Table)[0]
        rows = table.getElementsByType(TableRow)
        header = [str(c) for c in rows[0].childNodes]
        return header, len(rows) - 1
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return [], 0
    header = [str(c) for c in rows[0] if c is not None]
    body = [r for r in rows[1:] if any(c is not None for c in r)]
    return header, len(body)


@tool(
    name="verify_imports",
    description=(
        "Verify that the Excel workbook and the Google Sheet both contain the "
        "same data as the source CSV. Re-reads the saved workbook from disk and "
        "reads the Google Sheet back through the API, then compares row counts "
        "and headers. Call this last, before reporting to the user."
    ),
    input_schema={
        "type": "object",
        "properties": {
            "csv_path": {"type": "string", "description": "The source CSV."},
            "workbook_path": {"type": "string",
                              "description": "Workbook saved by import_csv_to_excel."},
            "spreadsheet_id": {"type": "string",
                               "description": "Spreadsheet written by import_csv_to_google_sheets."},
            "worksheet_title": {"type": "string", "default": "Employees"},
        },
        "required": ["csv_path"],
    },
)
def verify_imports(ctx: ToolContext, csv_path: str,
                   workbook_path: str | None = None,
                   spreadsheet_id: str | None = None,
                   worksheet_title: str | None = None) -> dict[str, Any]:
    source = Path(csv_path).expanduser()
    if not source.exists():
        return {"status": "failed", "error": f"CSV not found: {source}"}
    csv_header, csv_body = read_csv(source)
    checks: dict[str, Any] = {
        "csv": {"ok": True, "path": str(source), "rows": len(csv_body),
                "columns": len(csv_header)}
    }

    # --- Excel ------------------------------------------------------------
    wb_path = workbook_path or ctx.state.get("excel_path")
    if wb_path:
        path = Path(wb_path).expanduser()
        if not path.exists():
            checks["excel"] = {"ok": False, "error": f"workbook missing: {path}"}
        else:
            try:
                header, rows = _read_workbook(path)
                ok = rows == len(csv_body) and header == csv_header
                checks["excel"] = {
                    "ok": ok, "path": str(path), "rows": rows,
                    "expected_rows": len(csv_body),
                    "header_matches": header == csv_header,
                    "bytes": path.stat().st_size,
                }
            except Exception as exc:
                checks["excel"] = {"ok": False, "error": f"{type(exc).__name__}: {exc}"}
    else:
        checks["excel"] = {"ok": False, "error": "no workbook path known"}

    # --- Google Sheets ----------------------------------------------------
    sid = spreadsheet_id or ctx.state.get("spreadsheet_id")
    if sid:
        try:
            from .sheets_tool import build_service
            title = worksheet_title or ctx.config.sheets.worksheet_title
            service = build_service(ctx.config.sheets)
            values = service.spreadsheets().values().get(
                spreadsheetId=sid, range=f"'{title}'!A1:ZZ").execute().get("values", [])
            header = values[0] if values else []
            rows = max(len(values) - 1, 0)
            checks["google_sheets"] = {
                "ok": rows == len(csv_body) and header == csv_header,
                "spreadsheet_id": sid,
                "url": ctx.state.get("spreadsheet_url",
                                     f"https://docs.google.com/spreadsheets/d/{sid}"),
                "rows": rows, "expected_rows": len(csv_body),
                "header_matches": header == csv_header,
            }
        except Exception as exc:
            checks["google_sheets"] = {"ok": False,
                                       "error": f"{type(exc).__name__}: {exc}"}
    else:
        checks["google_sheets"] = {"ok": False, "error": "no spreadsheet id known"}

    all_ok = all(c.get("ok") for c in checks.values())
    return {
        "status": "success" if all_ok else "partial",
        "all_verified": all_ok,
        "checks": checks,
        "summary": ("Both imports verified against the CSV." if all_ok else
                    "Verification incomplete: " + ", ".join(
                        f"{k}={v.get('error', 'row/header mismatch')}"
                        for k, v in checks.items() if not v.get("ok"))),
    }
