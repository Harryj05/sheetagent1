"""Tool: import a CSV into Microsoft Excel and save the workbook.

Two engines:
  * ``com``      - drives the real Excel application through pywin32 (Windows).
                   This is what the assignment asks for: Excel visibly launches,
                   the CSV is opened as a workbook and saved as .xlsx.
  * ``openpyxl`` - headless fallback so the agent still completes on macOS,
                   Linux and CI. Produces the same workbook, without the GUI.

``engine: auto`` picks COM when it is importable and falls back otherwise,
reporting which engine actually ran so nothing is silently misrepresented.
"""
from __future__ import annotations

import csv
import logging
import platform
from pathlib import Path
from typing import Any

from ..registry import ToolContext, tool
from ..retry import with_retry

log = logging.getLogger("sheetagent.tools.excel")

SUPPORTED_FORMATS = {"xlsx", "xlsm", "csv", "ods"}
_XL_FORMATS = {"xlsx": 51, "xlsm": 52, "csv": 6, "ods": 60}


class ExcelUnavailable(RuntimeError):
    """Raised when the COM engine cannot be used at all - never retried.

    Semantically 'there is no Excel here'. With ``engine: auto`` this is what
    permits the openpyxl fallback.
    """


class ExcelPermanentError(RuntimeError):
    """A COM failure that will fail identically on every attempt.

    Semantically 'Excel is here and it refused'. Distinct from
    ``ExcelUnavailable`` because falling back to openpyxl would not help: a
    denied path or a bad argument breaks the headless writer too.
    """


# --------------------------------------------------------------------------- #
# COM failure classification
#
# The relative-path SaveAs bug is the motivating case: a deterministic error
# that burned three attempts and leaked an Excel process on each one. Two
# defences, in order of preference:
#
#   1. Reject the bad input before COM is touched at all (see the precondition
#      checks in _import_via_com). Cheaper and clearer than classifying the
#      resulting HRESULT.
#   2. Classify the HRESULT when Excel does reject something.
#
# Deliberately conservative: only codes whose meaning is unambiguous are listed.
# Anything unrecognised stays RETRYABLE - a wasted retry costs a few seconds,
# whereas wrongly marking a transient error permanent breaks a working run.
# --------------------------------------------------------------------------- #

#: Unambiguously deterministic - the same call will fail the same way.
_PERMANENT_HRESULTS = {
    -2147024894: "file not found (0x80070002)",
    -2147024891: "access denied (0x80070005)",
    -2147024809: "invalid argument (0x80070057)",
    -2147352571: "type mismatch in a COM argument (DISP_E_TYPEMISMATCH)",
    -2147352562: "wrong number of COM arguments (DISP_E_BADPARAMCOUNT)",
}

#: Unambiguously transient - Excel is busy, starting up, or holding a stale proxy.
_TRANSIENT_HRESULTS = {
    -2147418111: "call rejected by callee (RPC_E_CALL_REJECTED)",
    -2147417846: "server busy, retry later (RPC_E_SERVERCALL_RETRYLATER)",
    -2147417851: "server threw an exception (RPC_E_SERVERFAULT)",
    -2147023174: "RPC server unavailable (0x800706BA)",
    -2147023179: "RPC interface unknown - stale proxy (0x800706B5)",
    # A sharing violation is usually another process still holding the file -
    # frequently a previous Excel instance that is on its way out. Retrying is
    # exactly right here.
    -2147024864: "sharing violation, file locked (0x80070020)",
}

#: NOT classified on purpose. 0x800A03EC (-2146827284) is Excel's catch-all
#: "method of object failed" and covers both deterministic causes (a bad path)
#: and transient ones (a modal dialog open in another workbook). It is left
#: retryable; the relative-path case that motivated this work is caught by the
#: precondition check instead, which is where it belongs.


def _com_error_codes(exc: BaseException) -> tuple[int, ...]:
    """Extract the HRESULT and inner SCODE from a pywin32 ``com_error``.

    Matched by class name rather than ``isinstance``: ``pywintypes.com_error``
    only exists on Windows, and this must stay importable (and testable) on
    Linux CI.
    """
    if type(exc).__name__ != "com_error":
        return ()
    codes: list[int] = []
    args = getattr(exc, "args", ())
    if args and isinstance(args[0], int):
        codes.append(args[0])
    # args[2] is excepinfo; its 6th element is the SCODE Excel actually raised.
    if len(args) > 2 and isinstance(args[2], (tuple, list)) and len(args[2]) > 5             and isinstance(args[2][5], int):
        codes.append(args[2][5])
    return tuple(codes)


def _classify_com_error(exc: BaseException) -> BaseException:
    """Map a COM failure onto something ``with_retry`` can act on."""
    codes = _com_error_codes(exc)
    for code in codes:
        if code in _PERMANENT_HRESULTS:
            return ExcelPermanentError(
                f"Excel rejected the operation: {_PERMANENT_HRESULTS[code]}. "
                f"Retrying cannot help. Original error: {exc}")
    for code in codes:
        if code in _TRANSIENT_HRESULTS:
            log.debug("transient COM error %s: %s", code, _TRANSIENT_HRESULTS[code])
            return exc
    if codes:
        log.debug("unclassified COM error %s; treating as retryable", codes)
    return exc


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        raise ValueError(f"{path} is empty")
    return rows[0], rows[1:]


def _coerce(value: str) -> Any:
    """Excel should treat Salary as a number, not text."""
    text = value.strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return value


# --------------------------------------------------------------------------- #
# Engines
# --------------------------------------------------------------------------- #
def _import_via_com(csv_path: Path, out_path: Path, *, visible: bool,
                    close_after_save: bool, fmt: str) -> dict[str, Any]:
    # Excel resolves relative paths against ITS OWN working directory, not ours,
    # so a relative path here silently saves to the wrong place or fails with an
    # opaque "SaveAs method failed". The tool boundary normalises paths; this is
    # the assertion that the contract held. Deterministic, so it must never be
    # retried.
    for label, path in (("csv_path", csv_path), ("out_path", out_path)):
        if not path.is_absolute():
            raise ExcelPermanentError(
                f"{label} must be absolute when driving Excel via COM, got "
                f"{path!r}; Excel would resolve it against its own directory")
    if not csv_path.exists():
        raise ExcelPermanentError(
            f"CSV disappeared before Excel could open it: {csv_path}")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        import pythoncom  # type: ignore
        import win32com.client as win32  # type: ignore
    except ImportError as exc:  # pragma: no cover - Windows only
        raise ExcelUnavailable(
            "pywin32 is not installed / not on Windows; cannot drive Excel via COM"
        ) from exc


    pythoncom.CoInitialize()
    excel = None
    workbook = None
    succeeded = False
    try:
        try:
            excel = win32.Dispatch("Excel.Application")
        except Exception as exc:
            raise ExcelUnavailable(
                f"Microsoft Excel could not be started via COM: {exc}") from exc
        excel.Visible = visible
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(csv_path))
        if workbook is None:
            # Excel returns None instead of raising when it is mid-startup or a
            # modal dialog is blocking. Left retryable on purpose: observed to
            # succeed on the following attempt.
            raise RuntimeError(
                f"Excel returned no workbook when opening {csv_path} "
                "(Excel may be busy or showing a dialog)")
        sheet = workbook.Worksheets(1)
        sheet.Name = "Employees"
        used = sheet.UsedRange
        sheet.Rows(1).Font.Bold = True
        used.Columns.AutoFit()
        try:
            sheet.ListObjects.Add(1, used, None, 1).Name = "EmployeeTable"
        except Exception:  # table formatting is cosmetic, never fatal
            log.debug("could not add ListObject", exc_info=True)
        try:
            # Match the openpyxl path's frozen header. FreezePanes acts on the
            # window, not the sheet, so the workbook's own window is used
            # rather than ActiveWindow (which is None when Visible is False).
            window = workbook.Windows(1)
            window.SplitColumn = 0
            window.SplitRow = 1
            window.FreezePanes = True
        except Exception:
            log.debug("could not freeze header row", exc_info=True)
        rows = int(used.Rows.Count)
        cols = int(used.Columns.Count)
        version = str(excel.Version)
        workbook.SaveAs(str(out_path), FileFormat=_XL_FORMATS.get(fmt, 51))
        succeeded = True
        if close_after_save:
            workbook.Close(SaveChanges=False)
            excel.Quit()
        return {"engine": "com", "excel_version": version,
                "rows_in_sheet": rows, "columns_in_sheet": cols,
                "left_open": not close_after_save}
    finally:
        # A failed attempt must not leave an orphaned Excel process holding the
        # CSV open - the next retry would then fail for a different reason.
        if not succeeded:
            for close in (lambda: workbook.Close(SaveChanges=False),
                          lambda: excel.Quit()):
                try:
                    close()
                except Exception:
                    log.debug("cleanup after failed COM import", exc_info=True)
        pythoncom.CoUninitialize()


def _import_via_openpyxl(csv_path: Path, out_path: Path, fmt: str) -> dict[str, Any]:
    header, body = read_csv(csv_path)
    if fmt == "csv":
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerows([header, *body])
        return {"engine": "openpyxl", "rows_in_sheet": len(body) + 1,
                "columns_in_sheet": len(header)}
    if fmt == "ods":
        return _import_via_odf(header, body, out_path)

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employees"
    sheet.append(header)
    for row in body:
        sheet.append([_coerce(cell) for cell in row])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for idx, name in enumerate(header, start=1):
        width = max([len(name)] + [len(str(r[idx - 1])) for r in body if idx <= len(r)])
        sheet.column_dimensions[get_column_letter(idx)].width = min(width + 2, 40)
    ref = f"A1:{get_column_letter(len(header))}{len(body) + 1}"
    table = Table(displayName="EmployeeTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    try:
        sheet.add_table(table)
    except Exception:
        log.debug("could not add table", exc_info=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return {"engine": "openpyxl", "rows_in_sheet": len(body) + 1,
            "columns_in_sheet": len(header)}


def _import_via_odf(header: list[str], body: list[list[str]],
                    out_path: Path) -> dict[str, Any]:
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableCell, TableRow
        from odf.text import P
    except ImportError as exc:
        raise ExcelUnavailable("odfpy is required for .ods output (pip install odfpy)") from exc

    doc = OpenDocumentSpreadsheet()
    table = Table(name="Employees")
    for source in [header, *body]:
        tr = TableRow()
        for value in source:
            cell = TableCell(valuetype="string")
            cell.addElement(P(text=str(value)))
            tr.addElement(cell)
        table.addElement(tr)
    doc.spreadsheet.addElement(table)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(str(out_path))
    return {"engine": "odfpy", "rows_in_sheet": len(body) + 1,
            "columns_in_sheet": len(header)}


# --------------------------------------------------------------------------- #
# Tool
# --------------------------------------------------------------------------- #
@tool(
    name="import_csv_to_excel",
    description=(
        "Launch Microsoft Excel, import a CSV file into a workbook, format it and "
        "save the workbook to disk. On Windows this drives the real Excel "
        "application via COM; on other platforms it falls back to a headless "
        "openpyxl writer and says so in its result. Supports xlsx, xlsm, csv and ods."
    ),
    input_schema={
        "type": "object",
        "properties": {
            "csv_path": {"type": "string",
                         "description": "Path to the CSV produced by generate_employee_csv."},
            "output_path": {"type": "string",
                            "description": "Where to save the workbook. Defaults to the CSV name with the chosen extension."},
            "output_format": {"type": "string", "enum": sorted(SUPPORTED_FORMATS),
                              "default": "xlsx"},
            "visible": {"type": "boolean", "default": True,
                        "description": "Show the Excel window while working (COM engine only)."},
        },
        "required": ["csv_path"],
    },
)
def import_csv_to_excel(ctx: ToolContext, csv_path: str,
                        output_path: str | None = None,
                        output_format: str = "xlsx",
                        visible: bool | None = None) -> dict[str, Any]:
    cfg = ctx.config.excel
    fmt = output_format.lower().lstrip(".")
    if fmt not in SUPPORTED_FORMATS:
        return {"status": "failed",
                "error": f"unsupported output_format '{output_format}'",
                "supported": sorted(SUPPORTED_FORMATS)}

    source = Path(csv_path).expanduser()
    if not source.exists():
        return {"status": "failed", "error": f"CSV not found: {source}",
                "hint": "Run generate_employee_csv first."}

    # Normalise once, here, so both engines receive absolute paths and
    # _import_via_com's precondition check is a guard rather than a repair.
    source = source.resolve()
    target = (Path(output_path).expanduser() if output_path else
              Path(cfg.output_dir) / f"{source.stem}.{fmt}").resolve()

    engine = cfg.engine.lower()
    show = cfg.visible if visible is None else visible
    chosen = engine
    if engine == "auto":
        chosen = "com" if platform.system() == "Windows" else "openpyxl"

    ctx.events.emit("step_progress", f"Excel engine: {chosen}", engine=chosen)

    fallback_reason: str | None = None
    try:
        if chosen == "com":
            def _com_attempt() -> dict[str, Any]:
                try:
                    return _import_via_com(source, target, visible=show,
                                           close_after_save=cfg.close_after_save,
                                           fmt=fmt)
                except (ExcelUnavailable, ExcelPermanentError):
                    raise  # already classified
                except Exception as exc:
                    raise _classify_com_error(exc) from exc

            details = with_retry(
                _com_attempt,
                max_attempts=ctx.config.retry.max_attempts,
                initial_delay=ctx.config.retry.initial_delay,
                backoff=ctx.config.retry.backoff,
                max_delay=ctx.config.retry.max_delay,
                give_up_on=(ExcelUnavailable, ExcelPermanentError),
                label="excel-com-import",
            )
        else:
            details = _import_via_openpyxl(source, target, fmt)
    except ExcelPermanentError as exc:
        # Excel is present and refused. openpyxl would hit the same wall, so
        # substituting it would only hide the cause.
        return {"status": "failed", "error": str(exc), "engine_attempted": chosen,
                "retryable": False}
    except ExcelUnavailable as exc:
        if engine == "com":  # explicitly demanded COM - do not silently substitute
            return {"status": "failed", "error": str(exc),
                    "hint": "Set excel.engine to 'auto' or 'openpyxl' to run without Excel."}
        fallback_reason = str(exc)
        ctx.events.emit("step_progress", "Excel COM unavailable, using openpyxl",
                        reason=fallback_reason)
        details = _import_via_openpyxl(source, target, fmt)
    except Exception as exc:
        return {"status": "failed", "error": f"{type(exc).__name__}: {exc}",
                "engine_attempted": chosen}

    if not target.exists():
        return {"status": "failed",
                "error": f"engine reported success but {target} does not exist"}

    ctx.state["excel_path"] = str(target.resolve())
    result = {
        "status": "success",
        "workbook_path": str(target.resolve()),
        "bytes": target.stat().st_size,
        "format": fmt,
        **details,
        "summary": f"Saved workbook to {target} using the {details['engine']} engine",
    }
    if fallback_reason:
        result["fallback_reason"] = fallback_reason
        result["warning"] = ("Microsoft Excel was not available on this machine; "
                             "the workbook was written headlessly instead.")
    return result
