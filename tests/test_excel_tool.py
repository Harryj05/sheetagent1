import functools
import sys
from pathlib import Path

import pytest

from openpyxl import load_workbook

from sheetagent.tools.csv_tool import generate_employee_csv
from sheetagent.tools import excel_tool
from sheetagent.tools.excel_tool import import_csv_to_excel


def test_openpyxl_import_roundtrip(ctx):
    csv_result = generate_employee_csv(ctx=ctx, row_count=20, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert result["status"] == "success"
    assert result["engine"] == "openpyxl"

    workbook = load_workbook(result["workbook_path"])
    sheet = workbook["Employees"]
    assert sheet.max_row == 21
    assert sheet["A1"].value == "Employee ID"
    assert isinstance(sheet["E2"].value, int)      # Salary typed as a number
    assert sheet.freeze_panes == "A2"


def test_missing_csv_reports_failure(ctx):
    result = import_csv_to_excel(ctx=ctx, csv_path="/does/not/exist.csv")
    assert result["status"] == "failed"
    assert "not found" in result["error"]


def test_unsupported_format_rejected(ctx):
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"],
                                 output_format="numbers")
    assert result["status"] == "failed"
    assert "unsupported" in result["error"]


def test_csv_output_format(ctx):
    csv_result = generate_employee_csv(ctx=ctx, row_count=3, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"],
                                 output_format="csv",
                                 output_path=str(Path(ctx.config.output_dir) / "copy.csv"))
    assert result["status"] == "success"
    assert Path(result["workbook_path"]).exists()


def test_com_engine_failure_is_reported_not_swallowed(ctx, monkeypatch):
    """engine='com' must never silently substitute openpyxl.

    pywin32 is present on Windows dev machines and absent in CI, so the
    unavailability is simulated rather than assumed: setting a module to None in
    sys.modules makes ``import`` raise ImportError.
    """
    monkeypatch.setitem(sys.modules, "pythoncom", None)
    monkeypatch.setitem(sys.modules, "win32com.client", None)
    ctx.config.excel.engine = "com"
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert result["status"] == "failed"
    assert "COM" in result["error"] or "pywin32" in result["error"]
    assert not (Path(ctx.config.output_dir) / "employees.xlsx").exists()


def test_com_engine_used_when_excel_is_available(ctx, monkeypatch):
    """engine='com' routes through the COM path and reports engine='com'."""
    calls = {}

    def fake_com(csv_path, out_path, *, visible, close_after_save, fmt):
        calls["visible"] = visible
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(b"fake-xlsx")
        return {"engine": "com", "excel_version": "16.0",
                "rows_in_sheet": 3, "columns_in_sheet": 5, "left_open": False}

    monkeypatch.setattr(excel_tool, "_import_via_com", fake_com)
    ctx.config.excel.engine = "com"
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert result["status"] == "success"
    assert result["engine"] == "com"
    assert result["excel_version"] == "16.0"
    assert "fallback_reason" not in result
    assert calls["visible"] is ctx.config.excel.visible


# --------------------------------------------------------------------------- #
# COM retry classification: deterministic failures must cost exactly one
# attempt, transient ones must exhaust the retry budget.
# --------------------------------------------------------------------------- #
class com_error(Exception):
    """Stand-in for pywintypes.com_error, which exists only on Windows.

    excel_tool matches it by class name for exactly this reason, so this double
    exercises the real classification path on any platform.
    """


def _com(hresult, scode=None):
    excepinfo = (0, "Microsoft Excel", "method failed", "", 0, scode)
    return com_error(hresult, "Exception occurred.", excepinfo, None)


def _count_attempts(ctx, monkeypatch, exc):
    attempts = []

    def boom(*a, **kw):
        attempts.append(1)
        raise exc

    monkeypatch.setattr(excel_tool, "_import_via_com", boom)
    monkeypatch.setattr(excel_tool, "with_retry",
                        functools.partial(excel_tool.with_retry, sleep=lambda _: None))
    ctx.config.excel.engine = "com"
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    return result, len(attempts)


@pytest.mark.parametrize("hresult", sorted(excel_tool._PERMANENT_HRESULTS))
def test_permanent_com_errors_are_not_retried(ctx, monkeypatch, hresult):
    result, attempts = _count_attempts(ctx, monkeypatch, _com(hresult))
    assert attempts == 1, "a deterministic COM failure must not be retried"
    assert result["status"] == "failed"
    assert result["retryable"] is False


@pytest.mark.parametrize("hresult", sorted(excel_tool._TRANSIENT_HRESULTS))
def test_transient_com_errors_are_retried(ctx, monkeypatch, hresult):
    result, attempts = _count_attempts(ctx, monkeypatch, _com(hresult))
    assert attempts == ctx.config.retry.max_attempts
    assert result["status"] == "failed"


def test_permanent_scode_inside_excepinfo_is_classified(ctx, monkeypatch):
    """The real SCODE often hides in excepinfo behind DISP_E_EXCEPTION."""
    result, attempts = _count_attempts(
        ctx, monkeypatch, _com(-2147352567, scode=-2147024891))  # access denied
    assert attempts == 1
    assert result["retryable"] is False


def test_unclassified_com_errors_stay_retryable(ctx, monkeypatch):
    """0x800A03EC is Excel's catch-all; guessing at it would be wrong."""
    result, attempts = _count_attempts(
        ctx, monkeypatch, _com(-2147352567, scode=-2146827284))
    assert attempts == ctx.config.retry.max_attempts


def test_non_com_exceptions_stay_retryable(ctx, monkeypatch):
    result, attempts = _count_attempts(ctx, monkeypatch, RuntimeError("Excel busy"))
    assert attempts == ctx.config.retry.max_attempts


def test_relative_paths_rejected_before_com_is_touched(ctx, tmp_path):
    """The motivating bug: a relative SaveAs is deterministic, not transient."""
    with pytest.raises(excel_tool.ExcelPermanentError, match="must be absolute"):
        excel_tool._import_via_com(Path("output/employees.csv"),
                                   tmp_path / "out.xlsx",
                                   visible=False, close_after_save=True, fmt="xlsx")


def test_tool_boundary_hands_engines_absolute_paths(ctx, monkeypatch):
    seen = {}

    def capture(csv_path, out_path, **kw):
        seen["csv"], seen["out"] = csv_path, out_path
        out_path.write_bytes(b"x")
        return {"engine": "com", "rows_in_sheet": 3, "columns_in_sheet": 5}

    monkeypatch.setattr(excel_tool, "_import_via_com", capture)
    ctx.config.excel.engine = "com"
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert seen["csv"].is_absolute() and seen["out"].is_absolute()


def test_headless_runs_say_excel_was_not_launched(ctx):
    """A container/CI run must not read as though Excel had been driven."""
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert result["engine"] == "openpyxl"
    assert result["excel_launched"] is False
    assert "NOT launched" in result["warning"]


def test_com_runs_report_excel_was_launched(ctx, monkeypatch):
    monkeypatch.setattr(excel_tool, "_import_via_com",
                        lambda csv_path, out_path, **kw: (
                            out_path.write_bytes(b"x"),
                            {"engine": "com", "rows_in_sheet": 3,
                             "columns_in_sheet": 5})[1])
    ctx.config.excel.engine = "com"
    csv_result = generate_employee_csv(ctx=ctx, row_count=2, seed=1)
    result = import_csv_to_excel(ctx=ctx, csv_path=csv_result["csv_path"])
    assert result["excel_launched"] is True
    assert "warning" not in result
